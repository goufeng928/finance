import openpyxl  # OpenPyXL >= 3.0.7

# ##################################################

class PY3_xlsx_Replace_Content_Based_on_OpenPyXL3():

    def __init__(self):

        # openpyxl.load_workbook( ... ).worksheets[i] 输出示例:
        # >>> wb = openpyxl.load_workbook("example.xlsx")
        # >>> print(wb.worksheets)
        # [<Worksheet "Sheet1">,
        #  <Worksheet "Sheet2">,
        #  <Worksheet "Sheet3">]
        # >>> print(wb.worksheets[0])
        # <Worksheet "Sheet1">

        self.Column_Coordinate_Mapping = {
            # Microsoft Excel 2010 中列坐标最大值为 "XFD"
             0: "A",  1: "B",  2: "C",  3: "D",  4: "E",  5: "F",  6: "G",  7: "H",  8: "I",  9: "J",
            10: "K", 11: "L", 12: "M", 13: "N", 14: "O", 15: "P", 16: "Q", 17: "R", 18: "S", 19: "T",
            20: "U", 21: "V", 22: "W", 23: "X", 24: "Y", 25: "Z"
        }

    def xlsx_File_Input(self, xlsx_Path:str):

        wb = openpyxl.load_workbook(xlsx_Path)
        # ..........................................
        return wb

    def xlsx_Sheet_Replace_Image(self, xlsx_Sheet:list, Coord:str, Img_Path:str) -> object:

        # openpyxl.load_workbook( ... ).worksheets[i]._images[i].anchor._from 输出示例:
        # >>> wb = openpyxl.load_workbook("example.xlsx")
        # >>> print(wb.worksheets[0]._images[0].anchor._from)
        # <openpyxl.drawing.spreadsheet_drawing.AnchorMarker object>
        # Parameters:
        # col=6, colOff=45720, row=112, rowOff=38100

        # openpyxl.load_workbook( ... ).worksheets[i]._images[i].anchor._from.col 输出示例:
        # >>> wb = openpyxl.load_workbook("example.xlsx")
        # >>> print(wb.worksheets[0]._images[0].anchor._from.col)
        # 6

        # openpyxl.load_workbook( ... ).worksheets[i]._images[i].anchor._from.row 输出示例:
        # >>> wb = openpyxl.load_workbook("example.xlsx")
        # >>> print(wb.worksheets[0]._images[0].anchor._from.col)
        # 112

        xlsx_Sheet_Copy = xlsx_Sheet
        # ..........................................
        i:int = 0
        while (i < len(xlsx_Sheet_Copy._images)):
            xlsx_Img = xlsx_Sheet_Copy._images[i]
            # "_images" 是 openpyxl.load_workbook( ... ).worksheets[i] 的内部属性
            # ......................................
            if (isinstance(xlsx_Img.anchor, str) == True):
                xlsx_Img_Coord:str = xlsx_Img.anchor
                # ..................................
                if (xlsx_Img_Coord == Coord):
                    Bak_xlsx_Img_Anchor = xlsx_Img.anchor
                    Bak_xlsx_Img_Width  = xlsx_Img.width
                    Bak_xlsx_Img_Height = xlsx_Img.height
                    # ..............................
                    xlsx_Img_New = openpyxl.drawing.image.Image(Img_Path)
                    xlsx_Img_New.anchor = Bak_xlsx_Img_Anchor
                    xlsx_Img_New.width  = Bak_xlsx_Img_Width
                    xlsx_Img_New.height = Bak_xlsx_Img_Height
                    # ..............................
                    xlsx_Sheet_Copy._images[i] = xlsx_Img_New
            # ......................................
            if (isinstance(xlsx_Img.anchor, openpyxl.drawing.spreadsheet_drawing.TwoCellAnchor) == True):
                xlsx_Img_Col       = self.Column_Coordinate_Mapping.get(xlsx_Img.anchor._from.col)
                xlsx_Img_Row       = xlsx_Img.anchor._from.row + 1  # Excel 行数从 1 开始, Python 索引从 0 开始
                xlsx_Img_Coord:str = "%s%s" % (xlsx_Img_Col, xlsx_Img_Row)
                # ..................................
                if (xlsx_Img_Coord == Coord):
                    Bak_xlsx_Img_Anchor = xlsx_Img.anchor
                    Bak_xlsx_Img_Width  = xlsx_Img.width
                    Bak_xlsx_Img_Height = xlsx_Img.height
                    # ..............................
                    xlsx_Img_New = openpyxl.drawing.image.Image(Img_Path)
                    xlsx_Img_New.anchor = Bak_xlsx_Img_Anchor
                    xlsx_Img_New.width  = Bak_xlsx_Img_Width
                    xlsx_Img_New.height = Bak_xlsx_Img_Height
                    # ..............................
                    xlsx_Sheet_Copy._images[i] = xlsx_Img_New
            # ......................................
            i = i + 1
        # ..........................................
        return xlsx_Sheet_Copy

# EOF Signed by GF.
